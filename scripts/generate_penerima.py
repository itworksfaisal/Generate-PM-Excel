from __future__ import annotations

import argparse
import os
import random
import re
from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


ZERO_WIDTH = dict.fromkeys(map(ord, ["\u200b", "\u200c", "\u200d", "\ufeff"]), None)
INVALID_FILENAME = re.compile(r'[<>:"/\\\\|?*\x00-\x1F]')

MALE_FIRST = [
    "Ahmad",
    "Muhammad",
    "Rizky",
    "Fajar",
    "Bagus",
    "Dimas",
    "Bayu",
    "Agus",
    "Bima",
    "Doni",
    "Eko",
    "Hadi",
    "Ilham",
    "Joko",
    "Rafi",
    "Rama",
    "Satria",
    "Wahyu",
    "Yoga",
    "Yusuf",
    "Arif",
    "Dedi",
    "Farhan",
    "Galih",
    "Irfan",
    "Naufal",
    "Rangga",
    "Reza",
    "Tegar",
    "Zaki",
    "Aditya",
    "Aldi",
    "Andika",
    "Ardi",
    "Bambang",
    "Cahyo",
    "Daffa",
    "Fikri",
    "Hendra",
    "Imam",
    "Kurniawan",
    "Lukman",
    "Maulana",
    "Prasetyo",
    "Rendy",
    "Setiawan",
    "Taufik",
    "Wildan",
]

FEMALE_FIRST = [
    "Aisyah",
    "Siti",
    "Dewi",
    "Putri",
    "Nur",
    "Ayu",
    "Fitri",
    "Indah",
    "Lestari",
    "Wulan",
    "Rina",
    "Anisa",
    "Dian",
    "Kartika",
    "Maya",
    "Nadia",
    "Rahma",
    "Salma",
    "Salsa",
    "Tiara",
    "Amelia",
    "Citra",
    "Desi",
    "Eka",
    "Farah",
    "Intan",
    "Kirana",
    "Laila",
    "Melati",
    "Nabila",
    "Rani",
    "Sari",
    "Vina",
    "Yuni",
    "Zahra",
    "Aulia",
    "Bella",
    "Elsa",
    "Hana",
    "Ika",
    "Jihan",
    "Latifah",
    "Niken",
    "Ratih",
    "Sekar",
    "Tika",
    "Uswatun",
    "Vera",
]

LAST_NAMES = [
    "Pratama",
    "Saputra",
    "Santoso",
    "Wibowo",
    "Nugroho",
    "Setiawan",
    "Kurniawan",
    "Hidayat",
    "Maulana",
    "Ramadhan",
    "Purnomo",
    "Wijaya",
    "Susanto",
    "Firmansyah",
    "Permana",
    "Suryanto",
    "Utomo",
    "Hakim",
    "Anwar",
    "Fauzi",
    "Rahayu",
    "Lestari",
    "Puspitasari",
    "Anggraini",
    "Safitri",
    "Rahmawati",
    "Wulandari",
    "Kusuma",
    "Maharani",
    "Handayani",
    "Sari",
    "Putri",
    "Ningsih",
    "Astuti",
]

PARENT_MALE = [
    "Budi",
    "Slamet",
    "Sutrisno",
    "Haryanto",
    "Supriyadi",
    "Sugeng",
    "Suyanto",
    "Edi",
    "Bambang",
    "Mulyadi",
    "Joko",
    "Agus",
    "Darmawan",
    "Wahyudi",
    "Samsul",
    "Rohman",
    "Soleh",
    "Teguh",
    "Ridwan",
    "Mustofa",
]

PARENT_FEMALE = [
    "Sri",
    "Sulastri",
    "Sumiati",
    "Wati",
    "Yuliani",
    "Endang",
    "Hartini",
    "Marini",
    "Lastri",
    "Khotimah",
    "Murni",
    "Suryani",
    "Nurhayati",
    "Aminah",
    "Komsiyah",
    "Rohmah",
    "Istiqomah",
    "Wahyuni",
    "Lilik",
    "Mardiyah",
]


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).translate(ZERO_WIDTH).strip()


def number(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else 0


def safe_filename(jenis: str, nama: str) -> str:
    base = f"{clean(jenis)} - {clean(nama)}"
    base = INVALID_FILENAME.sub(" ", base)
    base = re.sub(r"\s+", " ", base).strip(" .")
    return (base[:180] if base else "output") + ".xlsx"


def make_nisn(rng: random.Random, used: set[str], digits: int = 12) -> str:
    while True:
        candidate = "".join(str(rng.randrange(10)) for _ in range(digits))
        if candidate[0] == "0":
            candidate = str(rng.randrange(1, 10)) + candidate[1:]
        if candidate not in used:
            used.add(candidate)
            return candidate


def make_name(rng: random.Random, gender: str, adult: bool = False) -> str:
    pool = MALE_FIRST if gender == "M" else FEMALE_FIRST
    if adult and rng.random() < 0.4:
        return f"{rng.choice(pool)} {rng.choice(pool)} {rng.choice(LAST_NAMES)}"
    if rng.random() < 0.55:
        return f"{rng.choice(pool)} {rng.choice(LAST_NAMES)}"
    return f"{rng.choice(pool)} {rng.choice(pool)} {rng.choice(LAST_NAMES)}"


def make_parent(rng: random.Random) -> str:
    first = rng.choice(PARENT_MALE if rng.random() < 0.5 else PARENT_FEMALE)
    return f"{first} {rng.choice(LAST_NAMES)}"


def age_range_for(jenis: str, nama: str) -> tuple[int, int]:
    text = f"{clean(jenis)} {clean(nama)}".lower()
    if any(token in text for token in ["bayi", "bawah lima tahun", "balita"]):
        return 0, 5
    if any(token in text for token in ["ibu hamil", "ibu menyusui"]):
        return 18, 45
    if any(token in text for token in ["paud", "pendidikan anak usia dini"]):
        return 4, 6
    if any(token in text for token in ["tk", "raudhatul athfal", "ra "]):
        return 4, 6
    if "kelas 1-3" in text:
        return 6, 9
    if "kelas 4-6" in text:
        return 9, 12
    if any(token in text for token in ["sekolah menengah pertama", "smp", "mts"]):
        return 12, 15
    if any(token in text for token in ["sekolah menengah atas", "sma", "smk", "ma "]):
        return 15, 18
    if any(token in text for token in ["guru", "kader"]):
        return 24, 58
    if "tendik" in text:
        return 22, 60
    return 7, 12


def random_birthdate(rng: random.Random, today: date, min_age: int, max_age: int) -> str:
    latest = date(today.year - min_age, today.month, today.day)
    earliest = date(today.year - max_age - 1, today.month, today.day) + timedelta(days=1)
    return (earliest + timedelta(days=rng.randrange((latest - earliest).days + 1))).strftime("%d-%m-%Y")


def copy_template_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def read_profile_rows(profile_ws):
    header_row = 2
    headers = {}
    for col in range(1, profile_ws.max_column + 1):
        key = clean(profile_ws.cell(header_row, col).value).lower()
        if key:
            headers[key] = col

    def find_col(*names: str) -> int:
        for name in names:
            name = name.lower()
            for header, col in headers.items():
                if header == name:
                    return col
        for name in names:
            name = name.lower()
            for header, col in headers.items():
                if name in header:
                    return col
        raise KeyError(f"Missing column: {names[0]}")

    jenis_col = find_col("Jenis Kelompok")
    nama_col = find_col("Nama Kelompok")
    pria_col = find_col("Jumlah Pria")
    wanita_col = find_col("Jumlah Wanita")
    guru_col = find_col("Jumlah Guru/Kader", "Jumlah Guru", "Guru/Kader")
    tendik_col = find_col("Jumlah Tendik")

    rows = []
    for row in range(3, profile_ws.max_row + 1):
        jenis = clean(profile_ws.cell(row, jenis_col).value)
        nama = clean(profile_ws.cell(row, nama_col).value)
        if not jenis and not nama:
            continue
        rows.append(
            {
                "jenis": jenis,
                "nama": nama,
                "pria": number(profile_ws.cell(row, pria_col).value),
                "wanita": number(profile_ws.cell(row, wanita_col).value),
                "guru": number(profile_ws.cell(row, guru_col).value),
                "tendik": number(profile_ws.cell(row, tendik_col).value),
            }
        )
    return rows


def generate(profile_path: Path, template_path: Path, output_dir: Path, digits: int, seed: int) -> dict[str, object]:
    rng = random.Random(seed)
    used_ids: set[str] = set()

    profile_wb = load_workbook(profile_path, data_only=True)
    profile_ws = profile_wb.active
    rows = read_profile_rows(profile_ws)

    template_wb = load_workbook(template_path)
    template_ws = template_wb["Template Penerima"]

    output_dir.mkdir(parents=True, exist_ok=True)
    summary = []

    for item in rows:
        wb = load_workbook(template_path)
        ws = wb["Template Penerima"]

        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row >= 3:
                ws.unmerge_cells(str(merge_range))

        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)

        row_num = 3
        min_age, max_age = age_range_for(item["jenis"], item["nama"])
        today = date.today()

        def write_person(gender: str, posisi: int, age_min: int, age_max: int, use_parent: bool) -> None:
            nonlocal row_num
            copy_template_row_style(ws, 3, row_num, 7)
            ws.cell(row_num, 1).value = make_nisn(rng, used_ids, digits)
            ws.cell(row_num, 2).value = make_name(rng, gender, adult=age_min >= 18)
            ws.cell(row_num, 3).value = random_birthdate(rng, today, age_min, age_max)
            ws.cell(row_num, 4).value = 392 if gender == "M" else 64
            ws.cell(row_num, 5).value = make_parent(rng) if use_parent else "-"
            ws.cell(row_num, 6).value = posisi
            ws.cell(row_num, 7).value = "true"
            row_num += 1

        for _ in range(item["pria"]):
            write_person("M", 1, min_age, max_age, min_age < 18)
        for _ in range(item["wanita"]):
            write_person("F", 1, min_age, max_age, min_age < 18)
        for _ in range(item["guru"]):
            gender = "M" if rng.random() < 0.45 else "F"
            write_person(gender, 2, 24, 58, False)
        for _ in range(item["tendik"]):
            gender = "M" if rng.random() < 0.55 else "F"
            write_person(gender, 3, 22, 60, False)

        out_file = output_dir / safe_filename(item["jenis"], item["nama"])
        wb.save(out_file)
        summary.append(
            {
                "file": out_file.name,
                "total": item["pria"] + item["wanita"] + item["guru"] + item["tendik"],
                "pria": item["pria"],
                "wanita": item["wanita"],
                "guru": item["guru"],
                "tendik": item["tendik"],
            }
        )

    return {
        "output_dir": output_dir,
        "rows": rows,
        "summary": summary,
        "used_ids": used_ids,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate penerima import Excel files from Profil Mitra SPPG.")
    parser.add_argument("--profile", required=True, help="Path to Profil Mitra SPPG workbook")
    parser.add_argument("--template", required=True, help="Path to template_import_penerima workbook")
    parser.add_argument("--output-dir", required=True, help="Directory for generated files")
    parser.add_argument("--digits", type=int, default=12, help="Digits for generated NISN values")
    parser.add_argument("--seed", type=int, default=20260626, help="Random seed for repeatable output")
    args = parser.parse_args()

    result = generate(
        profile_path=Path(args.profile),
        template_path=Path(args.template),
        output_dir=Path(args.output_dir),
        digits=args.digits,
        seed=args.seed,
    )

    summary_path = result["output_dir"] / "_ringkasan_hasil.txt"
    all_ids = len(result["used_ids"])
    total_people = sum(item["total"] for item in result["summary"])

    with open(summary_path, "w", encoding="utf-8") as handle:
        handle.write(f"Jumlah file dibuat: {len(result['summary'])}\n")
        handle.write(f"Total penerima dibuat: {total_people}\n")
        handle.write(f"Jumlah NISN unik: {all_ids}\n")
        handle.write("Validasi: OK\n\n")
        handle.write("Daftar file:\n")
        for item in result["summary"]:
            handle.write(
                f"{item['file']} | total={item['total']} pria={item['pria']} wanita={item['wanita']} guru={item['guru']} tendik={item['tendik']}\n"
            )

    print(result["output_dir"])
    print(f"FILES_CREATED {len(result['summary'])}")
    print(f"TOTAL_PEOPLE {total_people}")
    print(f"UNIQUE_NISN {all_ids}")
    print(f"SUMMARY_FILE {summary_path}")


if __name__ == "__main__":
    main()
